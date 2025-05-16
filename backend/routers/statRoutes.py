from fastapi import APIRouter, Depends, HTTPException, Response
from dependencies import get_db
import models
from sqlalchemy.orm import Session
from typing import Annotated, List, Dict
from sqlalchemy import func
from pydantic import BaseModel
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from datetime import datetime, timezone
from openpyxl.styles import Font, Alignment
from io import BytesIO
import io
import pytz
import math
import zipfile

router = APIRouter(
    prefix="/stats",
    tags=["Student Statistics"]
)

db_dependency = Annotated[Session, Depends(get_db)]

@router.get("/average_pronunciation_score/{student_id}")
async def get_average_pronunciation_score(student_id: int, db: db_dependency):
    average_score = db.query(func.avg(models.AssessmentHistory.score)).filter(models.AssessmentHistory.student_id == student_id).scalar()
    if average_score is None:
        raise HTTPException(status_code=404, detail="Student not found or no pronunciation assessments taken.")
    return {"average_pronunciation_score": average_score}

@router.get("/average_comprehension_score/{student_id}")
async def get_average_comprehension_score(student_id: int, db: db_dependency):
    average_score = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).filter(models.ComprehensionAssessmentHistory.student_id == student_id).scalar()
    if average_score is None:
        raise HTTPException(status_code=404, detail="Student not found or no comprehension assessments taken.")
    return {"average_comprehension_score": average_score}

@router.get("/pronunciation_scores/{student_id}")
async def get_pronunciation_scores(student_id: int, db: db_dependency):
    scores = db.query(models.AssessmentHistory.score).filter(models.AssessmentHistory.student_id == student_id).all()
    if not scores:
        raise HTTPException(status_code=404, detail="Student not found or no pronunciation assessments taken.")
    return {"pronunciation_scores": [score[0] for score in scores]}

@router.get("/comprehension_scores/{student_id}")
async def get_comprehension_scores(student_id: int, db: db_dependency):
    scores = db.query(models.ComprehensionAssessmentHistory.score).filter(models.ComprehensionAssessmentHistory.student_id == student_id).all()
    if not scores:
        raise HTTPException(status_code=404, detail="Student not found or no comprehension assessments taken.")
    return {"comprehension_scores": [score[0] for score in scores]}

@router.get("/average_score_by_level/{level_id}")
async def get_average_score_by_level(level_id: int, db: db_dependency):
    pronunciation_average = db.query(func.avg(models.AssessmentHistory.score)).join(models.User, models.User.user_id == models.AssessmentHistory.student_id).filter(models.User.level == level_id).scalar()
    comprehension_average = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id).filter(models.User.level == level_id).scalar()

    return {"pronunciation_average": pronunciation_average, "comprehension_average": comprehension_average}

@router.get("/assessment_counts_by_level/{level_id}")
async def get_assessment_counts_by_level(level_id: int, db: db_dependency):
    pronunciation_count = db.query(models.AssessmentHistory).join(models.User, models.User.user_id == models.AssessmentHistory.student_id).filter(models.User.level == level_id).count()
    comprehension_count = db.query(models.ComprehensionAssessmentHistory).join(models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id).filter(models.User.level == level_id).count()

    return {"pronunciation_count": pronunciation_count, "comprehension_count": comprehension_count}

@router.get("/score_trends/{student_id}")
async def get_score_trends(student_id: int, db: db_dependency):
    pronunciation_query = db.query(
        models.AssessmentHistory.date_taken,
        models.AssessmentHistory.score
    ).filter(models.AssessmentHistory.student_id == student_id)

    comprehension_query = db.query(
        models.ComprehensionAssessmentHistory.date_taken,
        models.ComprehensionAssessmentHistory.score
    ).filter(models.ComprehensionAssessmentHistory.student_id == student_id)

    pronunciation_scores = pronunciation_query.order_by(models.AssessmentHistory.date_taken).all()
    comprehension_scores = comprehension_query.order_by(models.ComprehensionAssessmentHistory.date_taken).all()

    pronunciation_data = [{"date": date.isoformat(), "score": score} for date, score in pronunciation_scores]
    comprehension_data = [{"date": date.isoformat(), "score": score} for date, score in comprehension_scores]

    return {
        "pronunciation_scores": pronunciation_data,
        "comprehension_scores": comprehension_data
    }
    
@router.get("/level/{level_id}/completion_rate")
async def get_level_completion_rate(level_id: int, db: db_dependency):
    pronunciation_query = db.query(
        func.date_trunc("day", models.AssessmentHistory.date_taken).label("date"),
        func.count(models.AssessmentHistory.history_id).label("count")
    ).join(models.User, models.User.user_id == models.AssessmentHistory.student_id).filter(models.User.level == level_id)

    comprehension_query = db.query(
        func.date_trunc("day", models.ComprehensionAssessmentHistory.date_taken).label("date"),
        func.count(models.ComprehensionAssessmentHistory.history_id).label("count")
    ).join(models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id).filter(models.User.level == level_id)

    pronunciation_counts = pronunciation_query.group_by("date").order_by("date").all()
    comprehension_counts = comprehension_query.group_by("date").order_by("date").all()

    pronunciation_data = [{"date": date.isoformat(), "count": count} for date, count in pronunciation_counts]
    comprehension_data = [{"date": date.isoformat(), "count": count} for date, count in comprehension_counts]

    return {
        "pronunciation_completion": pronunciation_data,
        "comprehension_completion": comprehension_data
    }
    
@router.get("/section/{section_id}/completion_rate")
async def get_section_completion_rate(section_id: int, db: db_dependency):
    pronunciation_query = db.query(
        func.date_trunc("day", models.AssessmentHistory.date_taken).label("date"),
        func.count(models.AssessmentHistory.history_id).label("count")
    ).join(models.User, models.User.user_id == models.AssessmentHistory.student_id).filter(models.User.section_id == section_id)

    comprehension_query = db.query(
        func.date_trunc("day", models.ComprehensionAssessmentHistory.date_taken).label("date"),
        func.count(models.ComprehensionAssessmentHistory.history_id).label("count")
    ).join(models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id).filter(models.User.section_id == section_id)

    pronunciation_counts = pronunciation_query.group_by("date").order_by("date").all()
    comprehension_counts = comprehension_query.group_by("date").order_by("date").all()

    pronunciation_data = [{"date": date.isoformat(), "count": count} for date, count in pronunciation_counts]
    comprehension_data = [{"date": date.isoformat(), "count": count} for date, count in comprehension_counts]

    return {
        "pronunciation_completion": pronunciation_data,
        "comprehension_completion": comprehension_data
    }

@router.get("/level/{level_id}/score_trends")
async def get_level_score_trends(level_id: int, db: db_dependency):
    pronunciation_query = db.query(
        models.AssessmentHistory.date_taken,
        models.AssessmentHistory.score,
        models.User.section_id
    ).join(models.User, models.User.user_id == models.AssessmentHistory.student_id).filter(models.User.level == level_id)

    comprehension_query = db.query(
        models.ComprehensionAssessmentHistory.date_taken,
        models.ComprehensionAssessmentHistory.score,
        models.User.section_id
    ).join(models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id).filter(models.User.level == level_id)

    pronunciation_scores = pronunciation_query.order_by(models.AssessmentHistory.date_taken).all()
    comprehension_scores = comprehension_query.order_by(models.ComprehensionAssessmentHistory.date_taken).all()

    pronunciation_data = [
        {"date": date.isoformat(), "score": score, "section_id": section_id}
        for date, score, section_id in pronunciation_scores
    ]

    comprehension_data = [
        {"date": date.isoformat(), "score": score, "section_id": section_id}
        for date, score, section_id in comprehension_scores
    ]
    return {
        "pronunciation_scores": pronunciation_data,
        "comprehension_scores": comprehension_data
    }
    
@router.get("/average_score/section_gender")
async def get_average_score_section_gender(db: db_dependency):
    sections = db.query(models.Sections).all()
    results = []

    for section in sections:
        section_data = {
            "section_id": section.section_id,
            "section_name": section.section_name,
            "gender_data": []
        }

        genders = ["Male", "Female"]  # Assuming these are the gender values
        for gender in genders:
            pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                models.User, models.User.user_id == models.AssessmentHistory.student_id
            ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()

            comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
            ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()

            section_data["gender_data"].append({
                "gender": gender,
                "pronunciation_average": pronunciation_avg,
                "comprehension_average": comprehension_avg
            })

        results.append(section_data)

    return results

@router.get("/average_score/gender")
async def get_average_score_gender(db: db_dependency):
    genders = ["Male", "Female"]
    results = []

    for gender in genders:
        pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
            models.User, models.User.user_id == models.AssessmentHistory.student_id
        ).filter(models.User.gender == gender).scalar()

        comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
            models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
        ).filter(models.User.gender == gender).scalar()

        results.append({
            "gender": gender,
            "pronunciation_average": pronunciation_avg,
            "comprehension_average": comprehension_avg
        })

    # Total average (no gender separation)
    total_pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).scalar()
    total_comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).scalar()

    results.append({
        "gender": "Total",
        "pronunciation_average": total_pronunciation_avg,
        "comprehension_average": total_comprehension_avg
    })

    return results

@router.get("/student_count/level_section_gender")
async def get_student_count_level_section_gender(db: db_dependency):
    sections = db.query(models.Sections).all()
    levels = db.query(models.PronunciationAssessmentType).all() #Assuming type_id in PronunciationAssessmentType is the same as level in User
    genders = ["Male", "Female"]
    results = []

    for section in sections:
        section_data = {
            "section_id": section.section_id,
            "section_name": section.section_name,
            "level_data": []
        }

        for level in levels:
            level_data = {
                "level_id": level.type_id,
                "level_name": level.type_name,
                "gender_counts": []
            }

            for gender in genders:
                count = db.query(func.count(models.User.user_id)).filter(
                    models.User.section_id == section.section_id,
                    models.User.level == level.type_id,
                    models.User.gender == gender
                ).scalar()

                level_data["gender_counts"].append({
                    "gender": gender,
                    "count": count
                })

            section_data["level_data"].append(level_data)

        results.append(section_data)

    return results

@router.get("/students/gender-distribution")
async def get_students_grouped_by_gender(db: db_dependency):
    student_counts = (
        db.query(models.User.gender, func.count(models.User.user_id).label("total_students"))
        .filter(models.User.role == "student")
        .group_by(models.User.gender)
        .all()
    )

    return [{"gender": gender, "total_students": total_students} for gender, total_students in student_counts]


# @router.get("/download")
# async def get_average_score_gender_excel(db: db_dependency):
#     from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
#     from openpyxl.utils import get_column_letter

#     sections = db.query(models.Sections).all()
#     genders = ["Male", "Female", "Total"]
#     results = []

#     for section in sections:
#         for gender in genders:
#             if gender != "Total":
#                 student_count = db.query(func.count(models.User.user_id)).filter(
#                     models.User.section_id == section.section_id,
#                     models.User.gender == gender
#                 ).scalar()

#                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.AssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()

#                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()
#             else:
#                 student_count = db.query(func.count(models.User.user_id)).filter(
#                     models.User.section_id == section.section_id,
#                 ).scalar()

#                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.AssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id, ).scalar()

#                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id, ).scalar()

#             results.append({
#                 "section_name": section.section_name,
#                 "gender": gender,
#                 "student_count": student_count or 0,
#                 "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
#                 "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
#             })

#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "ReadSpeak Report"

#     # === Style helpers ===
#     title_font = Font(size=14, bold=True)
#     header_font = Font(size=12, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     center_align = Alignment(horizontal="center", vertical="center")
#     border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )

#     # === Title ===
#     sheet.merge_cells("A1:E1")
#     sheet["A1"] = "ReadSpeak System - Student Performance Report"
#     sheet["A1"].font = title_font
#     sheet["A1"].alignment = center_align

#     # === Subtitle ===
#     sheet.merge_cells("A2:E2")
#     generated_on = datetime.now(pytz.timezone("Asia/Manila")).strftime("%B %d, %Y – %I:%M %p")
#     sheet["A2"] = f"Generated on {generated_on}"
#     sheet["A2"].alignment = Alignment(horizontal="right")

#     # === Table Headers ===
#     headers = ["Section", "Gender", "Student Count", "Pronunciation Average", "Comprehension Average"]
#     sheet.append(headers)
#     for col_num, header in enumerate(headers, 1):
#         cell = sheet.cell(row=3, column=col_num)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_align
#         cell.border = border
#         column_letter = get_column_letter(col_num)
#         sheet.column_dimensions[column_letter].width = 24

#     # === Table Data ===
#     for i, row in enumerate(results):
#         data_row = [
#             row["section_name"],
#             row["gender"],
#             row["student_count"],
#             row["pronunciation_average"],
#             row["comprehension_average"]
#         ]
#         sheet.append(data_row)
#         for col_num in range(1, 6):
#             cell = sheet.cell(row=4 + i, column=col_num)
#             cell.alignment = center_align
#             cell.border = border

#     # === Chart Data Below Table ===
#     chart_data_start = len(results) + 9
#     # print(chart_data_start)
#     genders_chart = ["Male", "Female", "Total"]
#     results_chart = []

#     for gender in genders_chart:
#         if gender != "Total":
#             pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
#                 models.User, models.User.user_id == models.AssessmentHistory.student_id
#             ).filter(models.User.gender == gender).scalar()

#             comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
#                 models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
#             ).filter(models.User.gender == gender).scalar()
#         else:
#             pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).scalar()
#             comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).scalar()
#         results_chart.append({
#             "gender": gender,
#             "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
#             "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
#         })

#     # Chart headers
#     sheet[f"A{chart_data_start - 1}"] = "Gender"
#     sheet[f"B{chart_data_start - 1}"] = "Pronunciation Average"
#     sheet[f"C{chart_data_start - 1}"] = "Comprehension Average"

#     sheet.merge_cells(f"A{chart_data_start-3}:B{chart_data_start-3}")
#     sheet[f"A{chart_data_start-3}"] = "Grade 3 Average(%) Performance by Gender"
#     sheet[f"A{chart_data_start-3}"].font = title_font
#     # sheet[f"A{chart_data_start-3}"].alignment = center_align
    
#     # Chart data
#     for i, row in enumerate(results_chart):
#         sheet[f"A{chart_data_start + i}"] = row["gender"]
#         sheet[f"B{chart_data_start + i}"] = row["pronunciation_average"]
#         sheet[f"C{chart_data_start + i}"] = row["comprehension_average"]

#     # Create the bar chart
#     chart = BarChart()
#     chart.title = "Average(%) Performance by Gender\n"
#     # chart.y_axis.scaling.max = 100
#     # chart.y_axis.scaling.min = 0
#     # chart.dataLabels = DataLabelList(showVal=True)
#     # chart.legend.position = 'b'
#     chart.y_axis.delete = False
#     chart.x_axis.delete = False
#     chart.y_axis.scaling.max = 100 
#     chart.y_axis.majorUnit = 10
#     chart.y_axis.scaling.min = 0
#     chart.y_axis.majorGridlines = None
#     chart.x_axis.majorGridlines = None
#     chart.dataLabels = DataLabelList()
#     chart.dataLabels.showVal = True
#     chart.dataLabels.showSerName = False
#     chart.dataLabels.showCatName = False
#     chart.dataLabels.showLegendKey = False
#     chart.legend = Legend()
#     chart.legend.position = 'b'
#     chart.legend.overlay = False
#     chart.title.overlay = False
#     chart.overlap = -20 

#     data = Reference(sheet, min_col=2, min_row=chart_data_start - 1, max_col=3, max_row=chart_data_start + 2)
#     categories = Reference(sheet, min_col=1, min_row=chart_data_start, max_row=chart_data_start + 2)
#     chart.add_data(data, titles_from_data=True)
#     chart.set_categories(categories)
#     chart.series[0].graphicalProperties.solidFill = "FF0000"
#     chart.series[1].graphicalProperties.solidFill = "FFFF00"
#     sheet.add_chart(chart, f"A{chart_data_start-1}")
    
#     # === Charts per section ===
#     section_chart_start_row = chart_data_start + 18  # start below the first chart
#     chart_width = 4  # columns between each chart
#     chart_height_rows = 16
#     charts_per_row = 2
    
#     # Adding chart header
#     chart_header_row = section_chart_start_row - 2  # Row above the first chart
#     sheet.merge_cells(f"A{chart_header_row}:B{chart_header_row}")
#     chart_header_cell = sheet[f"A{chart_header_row}"]
#     chart_header_cell.value = "Per Section Performance"
#     chart_header_cell.font = title_font
#     # chart_header_cell.alignment = Alignment(horizontal="center", vertical="center")

#     for idx, section in enumerate(sections):
#         chart_row = section_chart_start_row + (idx // charts_per_row) * chart_height_rows
#         chart_col = (idx % charts_per_row) * chart_width + 1
#         chart_anchor = f"{get_column_letter(chart_col)}{chart_row}"

#         # Fetch per-gender averages
#         section_gender_averages = []
#         for gender in ["Male", "Female", "Total"]:
#             if gender != "Total":
#                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.AssessmentHistory.student_id
#                 ).filter(models.User.gender == gender, models.User.section_id == section.section_id).scalar()

#                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
#                 ).filter(models.User.gender == gender, models.User.section_id == section.section_id).scalar()
#             else:
#                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.AssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id).scalar()

#                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
#                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
#                 ).filter(models.User.section_id == section.section_id).scalar()

#             section_gender_averages.append({
#                 "gender": gender,
#                 "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
#                 "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
#             })

#         # Write data to cells
#         data_start_row = chart_row + 2
#         sheet[f"{get_column_letter(chart_col)}{data_start_row - 1}"] = "Gender"
#         sheet[f"{get_column_letter(chart_col + 1)}{data_start_row - 1}"] = "Pronunciation Average"
#         sheet[f"{get_column_letter(chart_col + 2)}{data_start_row - 1}"] = "Comprehension Average"

#         for i, row in enumerate(section_gender_averages):
#             sheet.cell(row=data_start_row + i, column=chart_col, value=row["gender"])
#             sheet.cell(row=data_start_row + i, column=chart_col + 1, value=row["pronunciation_average"])
#             sheet.cell(row=data_start_row + i, column=chart_col + 2, value=row["comprehension_average"])

#         # === Chart styling (matches first chart) ===
#         chart = BarChart()
#         chart.title = f"{section.section_name} Performance by Gender\n"
#         chart.y_axis.scaling.max = 100
#         chart.y_axis.scaling.min = 0
#         chart.y_axis.majorUnit = 10
#         chart.y_axis.majorGridlines = None
#         chart.x_axis.majorGridlines = None
#         chart.dataLabels = DataLabelList()
#         chart.dataLabels.showVal = True
#         chart.dataLabels.showSerName = False
#         chart.dataLabels.showCatName = False
#         chart.dataLabels.showLegendKey = False
#         chart.legend = Legend()
#         chart.legend.position = 'b'
#         chart.legend.overlay = False
#         chart.title.overlay = False
#         chart.overlap = -20
#         chart.y_axis.delete = False
#         chart.x_axis.delete = False

#         # Add data
#         data = Reference(sheet, min_col=chart_col + 1, min_row=data_start_row - 1, max_col=chart_col + 2, max_row=data_start_row + 2)
#         categories = Reference(sheet, min_col=chart_col, min_row=data_start_row, max_row=data_start_row + 2)
#         chart.add_data(data, titles_from_data=True)
#         chart.set_categories(categories)

#         # Bar colors
#         chart.series[0].graphicalProperties.solidFill = "FF0000"  # red for pronunciation
#         chart.series[1].graphicalProperties.solidFill = "FFFF00"  # yellow for comprehension

#         # Place chart
#         sheet.add_chart(chart, chart_anchor)

#     # Dimensions
#     charts_per_row = 2
#     chart_height = 16  # estimated vertical space used per chart
#     start_row = chart_data_start + 5  # space after first gender chart

#     # Calculate max vertical position from rows of charts
#     num_charts = len(sections)
#     rows_of_charts = math.ceil(num_charts / charts_per_row)
#     last_chart_row = start_row + (rows_of_charts * chart_height)

#     # Footer
#     footer_row = last_chart_row + 15
#     sheet.merge_cells(f"A{footer_row}:E{footer_row}")
#     footer_cell = sheet[f"A{footer_row}"]
#     footer_cell.value = f"Notice: This report was automatically generated by the ReadSpeak system."
#     footer_cell.font = Font(size=10, italic=True)
#     footer_cell.alignment = Alignment(horizontal="left")

#     # Save and respond
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     headers = {
#         'Content-Disposition': 'attachment; filename="ReadSpeak_Report.xlsx"'
#     }
#     return Response(content=output.read(), headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# openpyxl imports for Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Matplotlib imports for chart generation
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker # Import for axis formatting

def page_template(canvas, doc):
    """Adds a header and footer to each page of the PDF."""
    canvas.saveState()
    styles = getSampleStyleSheet()

    # Header
    header_style = ParagraphStyle(name='HeaderStyle', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
    generated_on = datetime.now(pytz.timezone("Asia/Manila")).strftime("%B %d, %Y – %I:%M %p")
    header_text = f"Generated on {generated_on}"
    header_paragraph = Paragraph(header_text, header_style)

    # Calculate header width and height
    header_width, header_height = header_paragraph.wrap(doc.width, doc.topMargin)

    # Draw the header string at the top right of the drawable area
    # Use canvas.drawRightString for right alignment convenience
    canvas.setFont('Helvetica', 9) # Set font for drawString
    canvas.drawRightString(doc.width + doc.leftMargin, doc.height + doc.bottomMargin + doc.topMargin - header_height, header_text)


    # Footer
    footer_style = ParagraphStyle(name='FooterStyle', parent=styles['Italic'], fontSize=8, alignment=TA_LEFT)
    footer_text = "Notice: This report was automatically generated by the ReadSpeak system."
    footer_paragraph = Paragraph(footer_text, footer_style)

    # Calculate footer width and height
    footer_width, footer_height = footer_paragraph.wrap(doc.width, doc.bottomMargin)

    # Draw the footer string at the bottom left of the drawable area
    # Use canvas.drawString for left alignment
    canvas.setFont('Helvetica-Oblique', 8) # Set font for drawString (italic)
    canvas.drawString(doc.leftMargin, doc.bottomMargin - footer_height, footer_text)


    canvas.restoreState()


@router.get("/download")
async def get_full_report_zip(db: db_dependency):
    # --- 1. Data Fetching ---
    # Replicate your existing data fetching logic here.
    # Ensure you fetch all the data needed for both the Excel table
    # and the charts (overall and per-section).

    sections = db.query(models.Sections).all()
    genders = ["Male", "Female", "Total"]
    results_table = [] # Data for the main table

    # Data for the main table (Section, Gender, Count, Averages)
    for section in sections:
        for gender in genders:
            if gender != "Total":
                student_count = db.query(func.count(models.User.user_id)).filter(
                    models.User.section_id == section.section_id,
                    models.User.gender == gender
                ).scalar()

                pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                    models.User, models.User.user_id == models.AssessmentHistory.student_id
                ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()

                comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                    models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
                ).filter(models.User.section_id == section.section_id, models.User.gender == gender).scalar()
            else:
                # Total for the section
                student_count = db.query(func.count(models.User.user_id)).filter(
                    models.User.section_id == section.section_id,
                ).scalar()

                pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                    models.User, models.User.user_id == models.AssessmentHistory.student_id
                ).filter(models.User.section_id == section.section_id, ).scalar()

                comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                    models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
                ).filter(models.User.section_id == section.section_id, ).scalar()

            results_table.append({
                "section_name": section.section_name,
                "gender": gender,
                "student_count": student_count or 0,
                "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
                "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
            })

    # Data for the overall gender average chart (across all sections)
    results_chart_overall = []
    for gender in genders:
        if gender != "Total":
            pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                models.User, models.User.user_id == models.AssessmentHistory.student_id
            ).filter(models.User.gender == gender).scalar()

            comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
            ).filter(models.User.gender == gender).scalar()
        else:
            # Overall total
            pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).scalar()
            comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).scalar()
        results_chart_overall.append({
            "gender": gender,
            "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
            "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
        })

    # Data for per-section charts (nested structure)
    results_charts_sections = {}
    for section in sections:
        section_gender_averages = []
        for gender in genders:
            if gender != "Total":
                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                     models.User, models.User.user_id == models.AssessmentHistory.student_id
                 ).filter(models.User.gender == gender, models.User.section_id == section.section_id).scalar()

                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
                 ).filter(models.User.gender == gender, models.User.section_id == section.section_id).scalar()
            else:
                 pronunciation_avg = db.query(func.avg(models.AssessmentHistory.score)).join(
                     models.User, models.User.user_id == models.AssessmentHistory.student_id
                 ).filter(models.User.section_id == section.section_id).scalar()

                 comprehension_avg = db.query(func.avg(models.ComprehensionAssessmentHistory.score)).join(
                     models.User, models.User.user_id == models.ComprehensionAssessmentHistory.student_id
                 ).filter(models.User.section_id == section.section_id).scalar()

            section_gender_averages.append({
                 "gender": gender,
                 "pronunciation_average": round(pronunciation_avg, 2) if pronunciation_avg is not None else 0,
                 "comprehension_average": round(comprehension_avg, 2) if comprehension_avg is not None else 0
             })
        results_charts_sections[section.section_name] = section_gender_averages


    # --- 2. Generate Excel File ---
    # Reuse your existing openpyxl logic here.
    # It's already set up to write to a sheet.
    # We just need to save the workbook to a BytesIO object.
    excel_output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ReadSpeak Report"

    # === Style helpers ===
    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # === Title ===
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "ReadSpeak System - Student Performance Report"
    sheet["A1"].font = title_font
    sheet["A1"].alignment = center_align

    # === Subtitle ===
    sheet.merge_cells("A2:E2")
    generated_on = datetime.now(pytz.timezone("Asia/Manila")).strftime("%B %d, %Y – %I:%M %p")
    sheet["A2"] = f"Generated on {generated_on}"
    sheet["A2"].alignment = Alignment(horizontal="right")

    # === Table Headers ===
    headers = ["Section", "Gender", "Student Count", "Pronunciation Average", "Comprehension Average"]
    sheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 24

    # === Table Data ===
    for i, row in enumerate(results_table):
        data_row = [
            row["section_name"],
            row["gender"],
            row["student_count"],
            row["pronunciation_average"],
            row["comprehension_average"]
        ]
        sheet.append(data_row)
        for col_num in range(1, 6):
            cell = sheet.cell(row=4 + i, column=col_num)
            cell.alignment = center_align
            cell.border = border

    # === Chart Data Below Table (for Excel chart) ===
    # This data is written to the sheet so Excel can create the chart from it
    chart_data_start_excel = len(results_table) + 9
    sheet[f"A{chart_data_start_excel - 1}"] = "Gender"
    sheet[f"B{chart_data_start_excel - 1}"] = "Pronunciation Average"
    sheet[f"C{chart_data_start_excel - 1}"] = "Comprehension Average"

    sheet.merge_cells(f"A{chart_data_start_excel-3}:B{chart_data_start_excel-3}")
    sheet[f"A{chart_data_start_excel-3}"] = "Grade 3 Average(%) Performance by Gender"
    sheet[f"A{chart_data_start_excel-3}"].font = title_font

    for i, row in enumerate(results_chart_overall):
        sheet[f"A{chart_data_start_excel + i}"] = row["gender"]
        sheet[f"B{chart_data_start_excel + i}"] = row["pronunciation_average"]
        sheet[f"C{chart_data_start_excel + i}"] = row["comprehension_average"]

    # Create the overall bar chart for Excel
    chart_excel_overall = BarChart()
    chart_excel_overall.title = "Average(%) Performance by Gender\n"
    chart_excel_overall.y_axis.scaling.max = 100
    chart_excel_overall.y_axis.majorUnit = 10
    chart_excel_overall.y_axis.scaling.min = 0
    chart_excel_overall.y_axis.majorGridlines = None
    chart_excel_overall.x_axis.majorGridlines = None
    chart_excel_overall.dataLabels = DataLabelList()
    chart_excel_overall.dataLabels.showVal = True
    chart_excel_overall.dataLabels.showSerName = False
    chart_excel_overall.dataLabels.showCatName = False
    chart_excel_overall.dataLabels.showLegendKey = False
    chart_excel_overall.legend = Legend()
    chart_excel_overall.legend.position = 'b'
    chart_excel_overall.legend.overlay = False
    chart_excel_overall.title.overlay = False
    chart_excel_overall.overlap = -20
    chart_excel_overall.y_axis.delete = False
    chart_excel_overall.x_axis.delete = False

    data_excel_overall = Reference(sheet, min_col=2, min_row=chart_data_start_excel - 1, max_col=3, max_row=chart_data_start_excel + 2)
    categories_excel_overall = Reference(sheet, min_col=1, min_row=chart_data_start_excel, max_row=chart_data_start_excel + 2)
    chart_excel_overall.add_data(data_excel_overall, titles_from_data=True)
    chart_excel_overall.set_categories(categories_excel_overall)
    chart_excel_overall.series[0].graphicalProperties.solidFill = "FF0000" # Red
    chart_excel_overall.series[1].graphicalProperties.solidFill = "FFFF00" # Yellow

    sheet.add_chart(chart_excel_overall, f"A{chart_data_start_excel -1}") # Place chart

    # === Charts per section (for Excel) ===
    section_chart_start_row_excel = chart_data_start_excel + 18 # Start below the first chart
    chart_width_excel = 4 # columns between each chart
    chart_height_rows_excel = 16
    charts_per_row_excel = 2

    chart_header_row_excel = section_chart_start_row_excel - 2
    sheet.merge_cells(f"A{chart_header_row_excel}:B{chart_header_row_excel}")
    chart_header_cell_excel = sheet[f"A{chart_header_row_excel}"]
    chart_header_cell_excel.value = "Per Section Performance"
    chart_header_cell_excel.font = title_font

    for idx, section in enumerate(sections):
        chart_row_excel = section_chart_start_row_excel + (idx // charts_per_row_excel) * chart_height_rows_excel
        chart_col_excel = (idx % charts_per_row_excel) * chart_width_excel + 1
        chart_anchor_excel = f"{get_column_letter(chart_col_excel)}{chart_row_excel}"

        # Get the pre-fetched data for this section's chart
        section_gender_averages_excel = results_charts_sections[section.section_name]

        # Write data to cells for the section chart (can be hidden)
        data_start_row_excel = chart_row_excel + 2
        sheet[f"{get_column_letter(chart_col_excel)}{data_start_row_excel - 1}"] = "Gender"
        sheet[f"{get_column_letter(chart_col_excel + 1)}{data_start_row_excel - 1}"] = "Pronunciation Average"
        sheet[f"{get_column_letter(chart_col_excel + 2)}{data_start_row_excel - 1}"] = "Comprehension Average"

        for i, row in enumerate(section_gender_averages_excel):
            sheet.cell(row=data_start_row_excel + i, column=chart_col_excel, value=row["gender"])
            sheet.cell(row=data_start_row_excel + i, column=chart_col_excel + 1, value=row["pronunciation_average"])
            sheet.cell(row=data_start_row_excel + i, column=chart_col_excel + 2, value=row["comprehension_average"])

        # Create the section bar chart for Excel
        section_chart_excel = BarChart()
        section_chart_excel.title = f"{section.section_name} Performance by Gender\n"
        section_chart_excel.y_axis.scaling.max = 100
        section_chart_excel.y_axis.scaling.min = 0
        section_chart_excel.y_axis.majorUnit = 10
        section_chart_excel.y_axis.majorGridlines = None
        section_chart_excel.x_axis.majorGridlines = None
        section_chart_excel.dataLabels = DataLabelList()
        section_chart_excel.dataLabels.showVal = True
        section_chart_excel.dataLabels.showSerName = False
        section_chart_excel.dataLabels.showCatName = False
        section_chart_excel.dataLabels.showLegendKey = False
        section_chart_excel.legend = Legend()
        section_chart_excel.legend.position = 'b'
        section_chart_excel.legend.overlay = False
        section_chart_excel.title.overlay = False
        section_chart_excel.overlap = -20
        section_chart_excel.y_axis.delete = False
        section_chart_excel.x_axis.delete = False

        data_excel_section = Reference(sheet, min_col=chart_col_excel + 1, min_row=data_start_row_excel - 1, max_col=chart_col_excel + 2, max_row=data_start_row_excel + 2)
        categories_excel_section = Reference(sheet, min_col=chart_col_excel, min_row=data_start_row_excel, max_row=data_start_row_excel + 2)
        section_chart_excel.add_data(data_excel_section, titles_from_data=True)
        section_chart_excel.set_categories(categories_excel_section)

        section_chart_excel.series[0].graphicalProperties.solidFill = "FF0000" # red
        section_chart_excel.series[1].graphicalProperties.solidFill = "FFFF00" # yellow

        sheet.add_chart(section_chart_excel, chart_anchor_excel)

    # === Footer for Excel ===
    # Estimate footer row based on the last section chart position
    charts_per_row_excel = 2
    chart_height_excel = 16
    start_row_excel = chart_data_start_excel + 4 # Adjusted based on where the first chart was placed
    num_charts_excel = len(sections)
    rows_of_charts_excel = math.ceil(num_charts_excel / charts_per_row_excel)
    last_chart_approx_row_excel = start_row_excel + (rows_of_charts_excel * chart_height_excel)

    footer_row_excel = last_chart_approx_row_excel + 15
    sheet.merge_cells(f"A{footer_row_excel}:E{footer_row_excel}")
    footer_cell_excel = sheet[f"A{footer_row_excel}"]
    footer_cell_excel.value = f"Notice: This report was automatically generated by the ReadSpeak system."
    footer_cell_excel.font = Font(size=10, italic=True)
    footer_cell_excel.alignment = Alignment(horizontal="left")


    workbook.save(excel_output)
    excel_output.seek(0) # Rewind the buffer

    # --- 3. Generate PDF Report ---
    pdf_output = io.BytesIO()
    # Use landscape for potentially wider tables/charts if needed, but letter portrait is standard
    doc = SimpleDocTemplate(pdf_output, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []

    # Styles for ReportLab
    styles = getSampleStyleSheet()
    # Custom styles
    title_style = ParagraphStyle(name='ReportTitle', parent=styles['h1'], fontSize=16, spaceAfter=14, alignment=TA_CENTER)
    header_style = ParagraphStyle(name='TableHeader', parent=styles['Normal'], fontSize=10, textColor=colors.white, backColor=colors.HexColor("#4F81BD"), alignment=TA_CENTER, bold=True, leading=14)
    # Style for table data cells - using Paragraph inside Table cells
    cell_style = ParagraphStyle(name='TableCell', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, leading=12)
    # Modified section title style to be centered
    section_title_style = ParagraphStyle(name='SectionTitle', parent=styles['h2'], fontSize=14, spaceBefore=20, spaceAfter=10, alignment=TA_CENTER)
    # Removed subtitle_style and footer_style as they are handled by header/footer functions now
    signature_label_style = ParagraphStyle(name='SignatureLabel', parent=styles['Normal'], fontSize=10, spaceBefore=30, alignment=TA_LEFT)
    # Style for the signature line (underscores)
    signature_line_style = ParagraphStyle(name='SignatureLine', parent=styles['Normal'], fontSize=10, spaceAfter=10, alignment=TA_LEFT)
    
    # Add Main Title
    elements.append(Paragraph("ReadSpeak System - Student Performance Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Add Main Table
    # Prepare data for ReportLab Table - each cell needs to be a flowable (like Paragraph)
    table_data_rl = [
        [Paragraph(header, header_style) for header in ["Section", "Gender", "Student Count", "Pronunciation Average", "Comprehension Average"]]
    ]
    for row in results_table:
        table_data_rl.append([
            Paragraph(str(row["section_name"]), cell_style),
            Paragraph(str(row["gender"]), cell_style),
            Paragraph(str(row["student_count"]), cell_style),
            Paragraph(str(row["pronunciation_average"]), cell_style),
            Paragraph(str(row["comprehension_average"]), cell_style),
        ])

    # Define column widths (adjust as needed to fit page width)
    # Total page width is 8.5 inches. Margins are 0.75 each, so usable width is 7 inches.
    col_widths_rl = [1.5*inch, 1*inch, 1*inch, 1.75*inch, 1.75*inch] # Adjusted widths to fit 7 inches

    table = Table(table_data_rl, colWidths=col_widths_rl)

    # Table Style (matching Excel as much as possible)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")), # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), # Center align all cells
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Middle vertical align
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white), # Data rows background (alternating could be added)
        ('GRID', (0, 0), (-1, -1), 1, colors.black), # Cell borders
        ('BOX', (0, 0), (-1, -1), 1, colors.black), # Outer border
    ])
    table.setStyle(table_style)
    elements.append(table)
    elements.append(Spacer(1, 0.5*inch))

    # Add Overall Performance Chart
    elements.append(Paragraph("Grade 3 Average(%) Performance by Gender", section_title_style))

    # Generate Matplotlib chart for overall performance
    # Use a size that fits well within the PDF page width (7 inches usable)
    plt.figure(figsize=(6, 4)) # Adjust figure size (width, height in inches)
    genders_plot = [row["gender"] for row in results_chart_overall]
    pronunciation_scores = [row["pronunciation_average"] for row in results_chart_overall]
    comprehension_scores = [row["comprehension_average"] for row in results_chart_overall]

    x = range(len(genders_plot))
    width = 0.35

    plt.bar([i - width/2 for i in x], pronunciation_scores, width, label='Pronunciation', color='red')
    plt.bar([i + width/2 for i in x], comprehension_scores, width, label='Comprehension', color='yellow')

    plt.ylabel('Average Score (%)')
    plt.title('Average(%) Performance by Gender', y=1.1)
    plt.xticks(x, genders_plot)
    plt.ylim(0, 100) # Match Excel y-axis limit
    # Use mticker to format y-axis ticks as integers if they are scores
    plt.gca().yaxis.set_major_locator(mticker.MultipleLocator(10)) # Match Excel major units
    plt.legend(loc='lower center', bbox_to_anchor=(0.5, -0.3), ncol=2) # Place legend below chart
    plt.grid(axis='y', linestyle='--', alpha=0.7) # Add horizontal grid lines

    # Add data labels
    for i in x:
        plt.text(i - width/2, pronunciation_scores[i] + 2, str(pronunciation_scores[i]), ha='center', va='bottom', fontsize=8)
        plt.text(i + width/2, comprehension_scores[i] + 2, str(comprehension_scores[i]), ha='center', va='bottom', fontsize=8)

    # Add space above the title (adjust top parameter)
    # Increased top margin slightly for more space
    plt.tight_layout() # Adjust layout to make space for legend and title


    # Save chart to BytesIO
    overall_chart_buffer = io.BytesIO()
    plt.savefig(overall_chart_buffer, format='png') # PNG is good for charts
    overall_chart_buffer.seek(0)
    plt.close() # Close the figure to free memory

    # Add chart image to PDF elements
    # Image width should be less than or equal to the usable page width (7 inches)
    overall_chart_img = Image(overall_chart_buffer, width=6*inch, height=4.5*inch) # Adjust image size
    elements.append(overall_chart_img)
    elements.append(Spacer(1, 0.5*inch))


    # Add Section Performance Charts
    elements.append(Paragraph("Per Section Performance", section_title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Arrange section charts in a grid (e.g., 2 per row)
    charts_per_row_pdf = 2
    chart_elements_row = [] # List to hold charts for the current row

    for idx, (section_name, section_data) in enumerate(results_charts_sections.items()):
        # Generate Matplotlib chart for the section
        # Figure size adjusted for 2 charts per row within 7 inches usable width
        # (7 inches / 2 charts) - some padding = ~3.5 inches per chart
        # Increased height slightly for better spacing
        plt.figure(figsize=(3.5, 3.2)) # Smaller figure size for section charts

        genders_plot_section = [row["gender"] for row in section_data]
        pronunciation_scores_section = [row["pronunciation_average"] for row in section_data]
        comprehension_scores_section = [row["comprehension_average"] for row in section_data]

        x_section = range(len(genders_plot_section))
        width_section = 0.35

        plt.bar([i - width_section/2 for i in x_section], pronunciation_scores_section, width_section, label='Pronunciation', color='red')
        plt.bar([i + width_section/2 for i in x_section], comprehension_scores_section, width_section, label='Comprehension', color='yellow')

        plt.ylabel('Avg Score (%)') # Shorter label
        plt.title(f'{section_name} Performance', y=1.1) # Section specific title
        plt.xticks(x_section, genders_plot_section)
        plt.ylim(0, 100)
        plt.gca().yaxis.set_major_locator(mticker.MultipleLocator(20)) # Fewer ticks for smaller chart
        plt.legend(loc='lower center', bbox_to_anchor=(0.5, -0.4), ncol=2, fontsize=8) # Smaller legend
        plt.grid(axis='y', linestyle='--', alpha=0.7)

        # Add data labels
        for i in x_section:
            plt.text(i - width_section/2, pronunciation_scores_section[i] + 2, str(pronunciation_scores_section[i]), ha='center', va='bottom', fontsize=7) # Smaller font
            plt.text(i + width_section/2, comprehension_scores_section[i] + 2, str(comprehension_scores_section[i]), ha='center', va='bottom', fontsize=7)

        # Add space above the title (adjust top parameter)
        # Increased top margin slightly for more space
        plt.tight_layout() # Adjust layout for legend and title


        # Save chart to BytesIO
        section_chart_buffer = io.BytesIO()
        plt.savefig(section_chart_buffer, format='png')
        section_chart_buffer.seek(0)
        plt.close()

        # Add chart image to the current row of chart elements
        # Image width and height adjusted for larger size and 2 charts per row + padding
        section_chart_img = Image(section_chart_buffer, width=3.3*inch, height=2.8*inch) # Adjust size for grid
        chart_elements_row.append(section_chart_img)

        # If we have reached the number of charts per row or it's the last chart, add the row to elements
        if (idx + 1) % charts_per_row_pdf == 0 or (idx + 1) == len(results_charts_sections):
            # Create a table to hold charts in a row
            # If there's only one chart in the last row, put it in a single-cell table
            if len(chart_elements_row) == 1:
                 chart_row_table = Table([chart_elements_row])
            else: # Two charts in the row
                # Define widths for 2 columns, leaving space for padding
                chart_row_table = Table([chart_elements_row], colWidths=[3.4*inch, 3.4*inch])

            # Add spacing between charts in the row (if more than one)
            if len(chart_elements_row) > 1:
                 chart_row_table.setStyle(TableStyle([('LEFTPADDING', (1,0), (-1,0), 0.2*inch)])) # Add padding to charts after the first

            elements.append(chart_row_table)
            elements.append(Spacer(1, 0.3*inch)) # Space between rows of charts
            chart_elements_row = [] # Reset for the next row

    # --- Add Signature Field ---
    elements.append(Spacer(1, 0.4*inch)) # Space before the signature field
    elements.append(Paragraph("Report reviewed by:", signature_label_style))
    # Add a Spacer between the label and the line
    elements.append(Spacer(1, 0.1 * inch)) # Added spacer here
    # Add a line for the signature (using underscores)
    elements.append(Paragraph("_________________________", signature_line_style)) # Adjust number of underscores for line length
    elements.append(Spacer(1, 0.2*inch)) # Space after the signature line

    # Build the PDF
    # Pass the combined page_template function to onFirstPage and onLaterPages
    doc.build(elements, onFirstPage=page_template, onLaterPages=page_template)
    pdf_output.seek(0) # Rewind the buffer

    # --- 4. Create ZIP Archive ---
    zip_output = io.BytesIO()
    with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Add the Excel file
        zf.writestr("ReadSpeak_Report.xlsx", excel_output.getvalue())
        # Add the PDF file
        zf.writestr("ReadSpeak_Report_Printable.pdf", pdf_output.getvalue())

    zip_output.seek(0) # Rewind the buffer

    # --- 5. Return ZIP as FastAPI Response ---
    headers = {
        'Content-Disposition': 'attachment; filename="ReadSpeak_Report.zip"'
    }
    return Response(content=zip_output.read(), headers=headers, media_type="application/zip")